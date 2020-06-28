import openpyxl


def count_day(file, number, string_f):
    temp_file = open(file, 'r')
    file_data2 = temp_file.read()
    j = 0
    count = []
    while j < int(number):
        count.append(file_data2.count(string_f[j]))
        j = j + 1
    temp_file.close()
    return count


# reading the strings which we want to count
no_of_strings = input("Enter the number of strings you want to count\n")
string = []
i = 0
while i < int(no_of_strings):
    string.append(input("Enter string " + str(i + 1) + "\n"))
    i = i + 1

# storing count day wise
counts = []
days = input("Enter the number of days\n")
i = 0
while i < int(days):
    file1 = "D:\\day_" + str(i + 1) + ".txt"
    c1 = count_day(file1, no_of_strings, string)
    counts.append(c1)
    i = i + 1
# writing data in xls file
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "day1"
i = 0
while i < int(days):
    count1 = counts[i]
    j = 0
    while j < len(string):
        c1 = sheet.cell(row=j + 1, column=1)
        c1.value = string[j]
        c2 = sheet.cell(row=j + 1, column=2)
        c2.value = count1[j]
        j = j + 1
    wb.create_sheet(index=i + 1, title="day" + str(i + 2))
    sheet = wb.get_sheet_by_name('day' + str(i + 2))
    i = i + 1

wb.save("D:\\counts.xlsx")



