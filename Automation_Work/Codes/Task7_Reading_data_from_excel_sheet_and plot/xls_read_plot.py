import matplotlib.pyplot as plt
import openpyxl

# reading data from xls file
string_counts = []
strings = []
path = "D:\\counts.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active
m_row = sheet.max_row
no_of_sheets = len(wb.sheetnames)
i = 0
while i < int(no_of_sheets)-1:
    sheet = wb.get_sheet_by_name('day' + str(i + 1))
    r1 = []
    r2 = []
    for j in range(1, m_row + 1):
        cell = sheet.cell(row=j, column=2)
        cell1 = sheet.cell(row=j, column=1)
        r1.append(cell.value)
        r2.append(cell1.value)
    i = i + 1
    string_counts.append(r1)
    strings.append(r2)
string = strings[0]

# transpose of string_counts matrix for plotting count of strings day wise
n = len(string_counts)
m = len(string_counts[0])
count_string = []
for i in range(0, m):
    c2 = []
    for j in range(0, n):
        c2.append(string_counts[j][i])
    count_string.append(c2)

# plotting
type_of_graph = input("Enter the type of graph you want: 1)Bar 2)Line 3)Dot\n")
day_plot = []
k = 0
while k < int(no_of_sheets)-1:
    day_plot.append(k+1)
    k = k+1
i = 0
if type_of_graph == "Bar":
    for row in count_string:
        plt.bar(day_plot, row, label=(string[i]))
        i = i + 1
    plt.legend()
    plt.xlabel('DAYS')
    plt.ylabel('COUNT')
    plt.title('Count of Strings Day Wise')
    plt.show()

elif type_of_graph == "Line":
    for row in count_string:
        x = day_plot
        y = row
        plt.plot(x, y, label=(string[i]), linewidth=5)
        i = i + 1
    plt.title('Count of Strings Day Wise')
    plt.ylabel('COUNT')
    plt.xlabel('DAYS')
    plt.legend()
    plt.grid(True, color='k')
    plt.show()

else:
    for row in count_string:
        x = day_plot
        y = row
        plt.scatter(x, y, label=(string[i]))
        i = i + 1
    plt.xlabel('DAYS')
    plt.ylabel('COUNT')
    plt.title('Count of Strings Day Wise')
    plt.legend()
    plt.show()
