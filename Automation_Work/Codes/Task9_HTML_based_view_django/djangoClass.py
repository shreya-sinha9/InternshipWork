import matplotlib.pyplot as plt
import openpyxl


class SheetAccess:

    def __init__(self, days):
        self.days = days

    def plot_data(self, days):
        string_counts = []
        strings = []
        path = "D:\\counts.xlsx"
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        m_row = sheet.max_row
        i = 0
        while i < int(days):
            sheet = wb.get_sheet_by_name('day' + str(i + 1))
            r1 = []
            r2 = []
            for j in range(1, m_row + 1):
                cell = sheet.cell(row=j, column=2)
                cell1 = sheet.cell(row=j, column=1)
                r1.append(cell.value)
                r2.append(cell1.value)
            i = i + 1
            string_counts.append(r1)
            strings.append(r2)
        string = strings[0]

        # transpose of string_counts matrix for plotting count of strings day wise
        n = len(string_counts)
        m = len(string_counts[0])
        count_string = []
        for i in range(0, m):
            c2 = []
            for j in range(0, n):
                c2.append(string_counts[j][i])
            count_string.append(c2)
        return string, count_string


class PLOT:

    def __init__(self, days, type, count, string):
        self.days = days
        self.type = type
        self.count = count
        self.string = string

    def plot(self, days, type_of_graph, count_string, string):
        day_plot = []
        k = 0
        while k < int(days):
            day_plot.append(k + 1)
            k = k + 1
        i = 0
        if type_of_graph == "Bar":
            for row in count_string:
                plt.bar(day_plot, row, label=(string[i]))
                i = i + 1
            plt.legend()
            plt.xlabel('DAYS')
            plt.ylabel('COUNT')
            plt.title('Count of Strings Day Wise')
            plt.show()
            # plt.savefig('C:/Users/Shreya Sinha/try/media/temp4.png')

        elif type_of_graph == "Line":
            for row in count_string:
                x = day_plot
                y = row
                plt.plot(x, y, label=(string[i]), linewidth=5)
                i = i + 1
            plt.title('Count of Strings Day Wise')
            plt.ylabel('COUNT')
            plt.xlabel('DAYS')
            plt.legend()
            plt.grid(True, color='k')
            plt.show()
            # plt.savefig('C:/Users/Shreya Sinha/try/media/temp4.png')

        else:
            for row in count_string:
                x = day_plot
                y = row
                plt.scatter(x, y, label=(string[i]))
                i = i + 1
            plt.xlabel('DAYS')
            plt.ylabel('COUNT')
            plt.title('Count of Strings Day Wise')
            plt.legend()
            plt.show()
            # plt.savefig('C:/Users/Shreya Sinha/try/media/temp4.png')
