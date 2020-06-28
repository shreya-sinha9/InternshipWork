import matplotlib.pyplot as plt
import openpyxl
from tkinter import *

def plot_data(days):
    string_counts = []
    strings = []
    path = "D:\\counts.xlsx"
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    m_row = sheet.max_row
    i = 0
    while i < int(days):
        sheet = wb.get_sheet_by_name('day' + str(i + 1))
        r1 = []
        r2 = []
        for j in range(1, m_row + 1):
            cell = sheet.cell(row=j, column=2)
            cell1 = sheet.cell(row=j, column=1)
            r1.append(cell.value)
            r2.append(cell1.value)
        i = i + 1
        string_counts.append(r1)
        strings.append(r2)
    string = strings[0]

    # transpose of string_counts matrix for plotting count of strings day wise
    n = len(string_counts)
    m = len(string_counts[0])
    count_string = []
    for i in range(0, m):
        c2 = []
        for j in range(0, n):
            c2.append(string_counts[j][i])
        count_string.append(c2)
    return string,count_string

# plotting
def plot(type_of_graph, days, count_string, string):
    day_plot = []
    k = 0
    while k < int(days):
        day_plot.append(k + 1)
        k = k + 1
    i = 0
    if type_of_graph == "Bar":
        for row in count_string:
            plt.bar(day_plot, row, label=(string[i]))
            i = i + 1
        plt.legend()
        plt.xlabel('DAYS')
        plt.ylabel('COUNT')
        plt.title('Count of Strings Day Wise')
        plt.show()

    elif type_of_graph == "Line":
        for row in count_string:
            x = day_plot
            y = row
            plt.plot(x, y, label=(string[i]), linewidth=5)
            i = i + 1
        plt.title('Count of Strings Day Wise')
        plt.ylabel('COUNT')
        plt.xlabel('DAYS')
        plt.legend()
        plt.grid(True, color='k')
        plt.show()

    else:
        for row in count_string:
            x = day_plot
            y = row
            plt.scatter(x, y, label=(string[i]))
            i = i + 1
        plt.xlabel('DAYS')
        plt.ylabel('COUNT')
        plt.title('Count of Strings Day Wise')
        plt.legend()
        plt.show()

# GUI
root=Tk()
root.title('STRING FREQUENCY PLOTS(DAY WISE)')

def myClick():
   type_of_graph = clicked.get()
   days = E1.get()
   string,count_string = plot_data(days)
   plot(type_of_graph, days,count_string, string)

clicked =StringVar()
clicked.set("Line")
L1 = Label(root, text="Number of Days")
L1.grid(row=0,column=0)
E1=IntVar()
E1 = Entry(root, bd =5)
E1.grid(row=0,column=2)
L2 = Label(root, text="Type of graph")
L2.grid(row=1,column=0)
drop=OptionMenu(root,clicked,"Bar","Line","Dot")
drop.grid( row=1,column=2)

myButton=Button(root,text="Submit",command=myClick)
myButton.grid( row=2,column=1)

root.mainloop()